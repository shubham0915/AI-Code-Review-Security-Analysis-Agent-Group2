import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Software_Testing_Suite"

col_widths = [8, 38, 52, 42, 50, 18]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

HEADER_FILL  = PatternFill("solid", fgColor="1D5C3A")
ALT_FILL     = PatternFill("solid", fgColor="E8F5E9")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT    = Font(name="Calibri", size=10)
BOLD_FONT    = Font(name="Calibri", size=10, bold=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

thin = Side(style="thin", color="BDBDBD")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = HEADER_FILL; c.font = HEADER_FONT
    c.alignment = CENTER_ALIGN; c.border = BORDER

def cell(row, col, value, bold=False, fill=None, align=LEFT_ALIGN):
    c = ws.cell(row=row, column=col, value=value)
    c.font = BOLD_FONT if bold else BODY_FONT
    c.alignment = align; c.border = BORDER
    c.fill = fill if fill else (ALT_FILL if row % 2 == 0 else WHITE_FILL)

ws.row_dimensions[1].height = 32
headers = ["Sl. No", "Test Case Name", "Test Procedure Name",
           "Condition to be Tested", "Expected Result", "Actual Result"]
for col, h in enumerate(headers, 1):
    hdr(1, col, h)

tests = [
    (1,  "Valid Python Submission",
         "Pass simple Python arithmetic expression to validate_code()",
         "Syntactically correct Python code",
         "valid=True, errors=[]", "Passed"),
    (2,  "Valid Python Function",
         "Pass a full Python function with type hints to validate_code()",
         "Well-formed function with return type annotation",
         "valid=True returned", "Passed"),
    (3,  "Invalid Python Syntax",
         "Submit broken Python code with unclosed parenthesis",
         "SyntaxError present in code",
         "valid=False, SyntaxError in errors[0].message", "Passed"),
    (4,  "Invalid Python Indentation",
         "Submit Python function with misaligned return statement",
         "IndentationError present",
         "valid=False", "Passed"),
    (5,  "Empty Code Submission",
         "Submit empty string to validate_code()",
         "Empty string input",
         "valid=False, 'empty' in detail message", "Passed"),
    (6,  "Whitespace-Only Code",
         "Submit string of spaces, tabs, and newlines to validate_code()",
         "No actual code tokens present",
         "valid=False", "Passed"),
    (7,  "Python with Imports",
         "Submit valid Python module with os import and function",
         "Valid multi-line Python with stdlib imports",
         "valid=True", "Passed"),
    (8,  "Valid Java Simple Class",
         "Submit minimal Java HelloWorld class to validate_code()",
         "Syntactically correct Java class with main method",
         "valid=True, errors=[]", "Passed"),
    (9,  "Invalid Java — No Class Keyword",
         "Submit Java code missing class declaration",
         "No 'class' keyword present",
         "valid=False, 'class' in error message", "Passed"),
    (10, "Invalid Java — Unbalanced Braces",
         "Submit Java code with missing closing brace for class body",
         "Unbalanced { } braces",
         "valid=False, 'brace' in error message", "Passed"),
    (11, "Valid Java with Imports",
         "Submit Java class with List and ArrayList imports",
         "Valid multi-import Java class definition",
         "valid=True", "Passed"),
    (12, "Detect Python by File Extension",
         "Call detect_language() with .py filename",
         "Filename suffix is .py",
         "Language.python returned", "Passed"),
    (13, "Detect Java by File Extension",
         "Call detect_language() with .java filename",
         "Filename suffix is .java",
         "Language.java returned", "Passed"),
    (14, "Auto-Detect Python by Keywords",
         "Call detect_language() with def/print keywords, no filename",
         "Code contains Python-specific keywords",
         "Language.python returned", "Passed"),
    (15, "Auto-Detect Java by Keywords",
         "Call detect_language() with public class/main keywords, no filename",
         "Code contains Java-specific keywords",
         "Language.java returned", "Passed"),
    (16, "Bandit Security Linter — Python",
         "Run run_bandit() on temp Python file, capture JSON output",
         "Valid Python file path provided",
         "Dict returned with 'results' key from Bandit", "Passed"),
    (17, "Pylint Code Quality Linter",
         "Run run_pylint() on temp Python file",
         "Valid Python file with lintable code",
         "List of Pylint message objects returned", "Passed"),
    (18, "Radon Complexity Metrics",
         "Run run_radon() on temp Python file",
         "Valid Python file provided",
         "Dict with 'cc' and 'raw' keys returned", "Passed"),
    (19, "Python Full Linter Pipeline",
         "Call run_python_linters() with vulnerable Python code string",
         "Valid Python code submitted",
         "Dict with 'bandit', 'pylint', 'radon' keys returned", "Passed"),
    (20, "Semgrep Java Linter — Structure",
         "Call run_java_linters() with minimal valid Java class",
         "Semgrep binary available in .venv/bin/",
         "Dict with 'heuristics' (list) and 'semgrep' (dict) keys", "Passed"),
    (21, "Guardrail — Valid Code Accepted",
         "Call validate_intent() with a real Python function",
         "LLM classifies input as valid source code",
         "is_valid=True, reason contains 'Valid'", "Passed"),
    (22, "Guardrail — Prompt Injection Rejected",
         "Call validate_intent() with jailbreak prompt text",
         "LLM classifies input as prompt injection",
         "is_valid=False, reason contains 'prompt injection'", "Passed"),
    (23, "Guardrail — Too Short Input",
         "Call validate_intent() with 3-character string 'def'",
         "Input length below minimum threshold",
         "is_valid=False, 'too short' in reason (no LLM call)", "Passed"),
    (24, "Code Analysis Agent — Success Path",
         "Call run_code_analysis() with mocked LLM returning valid result",
         "Mocked LLM returns CodeAnalysisResult score=95, grade='A'",
         "Result contains CodeAnalysisResult with quality_score=95", "Passed"),
    (25, "Code Analysis Agent — JSON Parse Fallback",
         "Call run_code_analysis() with LLM returning invalid JSON twice",
         "LLM returns non-JSON string on both retry attempts",
         "Fallback: quality_score=0, grade='F', '[PARSE ERROR]' in summary", "Passed"),
    (26, "Security Vulnerability Agent — Success",
         "Call run_security_vuln() with mocked LLM returning zero-vuln JSON",
         "LLM returns valid SecurityAnalysisResult JSON, score=100",
         "SecurityAnalysisResult with security_score=100 returned", "Passed"),
    (27, "Remediation Agent — With Findings",
         "Call run_remediation() with CodeSmell ca-001 injected in state",
         "State contains CodeSmell with SQL injection description",
         "RemediationResult with 'parameterized' in summary", "Passed"),
    (28, "Remediation Agent — No Findings (Early Return)",
         "Call run_remediation() with clean state and no findings",
         "State has no findings in any result field",
         "RemediationResult with empty remediations and 'No findings'", "Passed"),
    (29, "PR Summary Agent — Success Path",
         "Call run_pr_summary() with mocked LLM returning HIGH risk JSON",
         "LLM returns PRSummaryResult overall_risk=HIGH, approved=False",
         "PRSummaryResult: overall_risk=HIGH, approved=False, security=65", "Passed"),
    (30, "PR Summary Agent — LLM Fallback",
         "Call run_pr_summary() with LLM returning non-JSON twice",
         "LLM fails on both retry attempts; deterministic fallback triggered",
         "Scores from state fixture: security=65, quality=70", "Passed"),
    (31, "Fallback Scores — Clean State",
         "Call _compute_fallback_scores() with empty findings state",
         "State has no vulnerabilities or code smells",
         "overall_risk='CLEAN', total_findings=0, approved=True", "Passed"),
    (32, "Fallback Scores — High Risk State",
         "Call _compute_fallback_scores() with state having high-severity vuln",
         "State security result has high_count=1",
         "overall_risk='HIGH', approved=False", "Passed"),
    (33, "Collect Findings — Code Smells",
         "Call _collect_findings() on state with injected ca-001 CodeSmell",
         "CodeSmell with id='ca-001' in code_analysis_result.findings",
         "List contains item with id='ca-001'", "Passed"),
    (34, "RAG Query Builder — Security Priority",
         "Call _build_rag_query() with mixed security + code smell findings",
         "Findings list has security_vulnerability type as first item",
         "Query string contains 'SQL Injection' and 'remediation'", "Passed"),
    (35, "Chat Graph Initialization",
         "Instantiate chat graph and verify checkpointer configuration",
         "MemorySaver checkpoint store initialized",
         "Graph compiles without error; checkpointer attached", "Passed"),
    (36, "Chat Graph with Session Context",
         "Send follow-up question with pre-cached analysis result in session",
         "Session result injected in cache; multi-turn conversation invoked",
         "Response references previously analysed code; memory preserved", "Passed"),
    (37, "API Health Check",
         "Send GET /health to FastAPI application",
         "Server is running and reachable",
         "HTTP 200 with {status: 'ok'}", "Passed"),
    (38, "API Submit Valid Python (Paste)",
         "POST /api/v1/submit/paste with valid Python code payload",
         "Redis mocked; valid Python payload with language='python'",
         "HTTP 202, session_id present, status='queued', language='python'", "Passed"),
    (39, "API Submit Valid Java (Paste)",
         "POST /api/v1/submit/paste with valid Java class code",
         "Redis mocked; valid Java class payload sent",
         "HTTP 202, language='java' in response body", "Passed"),
    (40, "API Auto-Detects Python Language",
         "POST /api/v1/submit/paste with language='auto' and Python code",
         "Auto-detection mode selected by client",
         "HTTP 202, language='python' correctly detected and returned", "Passed"),
    (41, "API Rejects Broken Python Syntax",
         "POST /api/v1/submit/paste with Python code containing SyntaxError",
         "Code with unclosed parenthesis submitted",
         "HTTP 422 Unprocessable Entity returned", "Passed"),
    (42, "API Rejects Empty Code Body",
         "POST /api/v1/submit/paste with empty code string",
         "Empty string fails Pydantic min_length=1 constraint",
         "HTTP 422 returned", "Passed"),
    (43, "API Rejects Oversized Code (>10K lines)",
         "POST /api/v1/submit/paste with 11,000-line Python code block",
         "Code exceeds maximum allowed line count limit",
         "HTTP 413 Request Entity Too Large returned", "Passed"),
    (44, "API Validate Endpoint — Valid Python",
         "POST /api/v1/submit/validate with syntactically valid Python",
         "Valid Python code submitted to validate-only endpoint",
         "HTTP 200, valid=True in response JSON", "Passed"),
    (45, "API Validate Endpoint — Invalid Python",
         "POST /api/v1/submit/validate with broken Python code",
         "Code with syntax error submitted to validate endpoint",
         "HTTP 200, valid=False, errors list is non-empty", "Passed"),
    (46, "API Validate Endpoint — Valid Java",
         "POST /api/v1/submit/validate with minimal valid Java class",
         "Correct Java class body submitted to validate endpoint",
         "HTTP 200, valid=True in response", "Passed"),
    (47, "API File Upload — Valid Python .py File",
         "POST /api/v1/submit/file with multipart .py file upload",
         "Valid Python file with correct .py extension",
         "HTTP 202, language='python' in response body", "Passed"),
    (48, "API File Upload — Unsupported Extension",
         "POST /api/v1/submit/file with .cpp C++ source file",
         "Unsupported file extension (.cpp) submitted",
         "HTTP 415 Unsupported Media Type returned", "Passed"),
    (49, "Streamlit App Loads in Local Mode",
         "Import frontend/app.py and verify LOCAL_MODE flag is True",
         "Streamlit app loaded without live API server connection",
         "LOCAL_MODE=True; no connection errors raised", "Passed"),
    (50, "Frontend Local Validation Function",
         "Invoke local_validate() directly with invalid Python code",
         "SyntaxError code passed to frontend local validator",
         "Validation error message string returned (non-empty)", "Passed"),
    (51, "React UI — Severity Normalization",
         "Mount ResultsPanel.tsx with mock finding containing 'CRITICAL' severity",
         "normalizeSeverity() parses raw severity string",
         "Finding binned into the 'High / Critical' top tier visual group", "Passed"),
    (52, "React UI — PR Summary Fallbacks",
         "Mount PRSummaryPanel.tsx with missing release_notes array in payload",
         "Array.isArray() checks execute before .map()",
         "Component gracefully renders 'No release notes generated.' without crashing", "Passed"),
    (53, "React UI — PDF Export Button",
         "Click 'Download Full PDF Report' in Results Panel",
         "react-to-print hook invoked targeting the reportRef DOM node",
         "Browser native print dialog opens; @media print CSS rules perfectly retain dark mode", "Passed"),
    (54, "Celery Concurrency — macOS Apple Silicon",
         "Start Celery worker with --pool=solo flag and submit LangGraph request",
         "Parallel graph execution via asyncio occurs inside the single solo worker process",
         "Task completes successfully without SIGSEGV fork() collisions", "Passed"),
]

for row_idx, (sl, name, proc, cond, expected, actual) in enumerate(tests, 2):
    ws.row_dimensions[row_idx].height = 46
    row_fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
    cell(row_idx, 1, sl,       bold=True, fill=row_fill, align=CENTER_ALIGN)
    cell(row_idx, 2, name,     fill=row_fill)
    cell(row_idx, 3, proc,     fill=row_fill)
    cell(row_idx, 4, cond,     fill=row_fill)
    cell(row_idx, 5, expected, fill=row_fill)
    c = ws.cell(row=row_idx, column=6, value=actual)
    c.font  = Font(name="Calibri", size=10, bold=True,
                   color="2E7D32" if actual == "Passed" else "C62828")
    c.alignment = CENTER_ALIGN; c.border = BORDER; c.fill = row_fill

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:F{len(tests)+1}"

out = "/Users/arous/Desktop/AI-Code-Review-Security-Analysis-Agent-Group2/docs/Unit_Test_Plan_v0.1.xlsx"
wb.save(out)
print(f"Saved -> {out}  ({len(tests)} test cases)")
