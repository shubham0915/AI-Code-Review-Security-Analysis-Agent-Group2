"""
Generate Defect_Tracker_Filled_v0.1.xlsx for the AI Code Review Security Analysis Agent project.
All defects sourced from documented debugging sessions (DEBUGGING_SESSION_NOTES.md Sections 1-11).
Columns: Sl No | Submitted By | Submitted Date | Description | Detected Sprint |
         Assigned To | Type Of Defect | Action Taken | Action Taken Date | Status | Remarks
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load the original template to preserve sheet styles/validation
wb = openpyxl.load_workbook(
    "/Users/arous/Desktop/AI-Code-Review-Security-Analysis-Agent-Group2/docs/Defect_Tracker Template_v0.1.xlsx"
)
ws = wb["Defects"]

# ── Styles ─────────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1D5C3A")
ALT1        = PatternFill("solid", fgColor="E8F5E9")
ALT2        = PatternFill("solid", fgColor="FFFFFF")
CLOSED_FILL = PatternFill("solid", fgColor="E8F5E9")
OPEN_FILL   = PatternFill("solid", fgColor="FFEBEE")

HDR_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT  = Font(name="Calibri", size=10)
BOLD_FONT  = Font(name="Calibri", size=10, bold=True)
CLOSED_FNT = Font(name="Calibri", size=10, bold=True, color="2E7D32")
OPEN_FNT   = Font(name="Calibri", size=10, bold=True, color="C62828")

C = Alignment(horizontal="center", vertical="center", wrap_text=True)
L = Alignment(horizontal="left",   vertical="center", wrap_text=True)

thin  = Side(style="thin",   color="BDBDBD")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Column widths ──────────────────────────────────────────────────────────────
col_widths = [7, 14, 14, 72, 16, 14, 20, 68, 16, 18, 44]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Re-style header row ────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 32
headers = ["Sl No", "Submitted By", "Submitted Date", "Description",
           "Detected Sprint", "Assigned To", "Type Of Defect",
           "Action Taken", "Action Taken Date", "Status\n(Open/Closed)", "Remarks"]
for ci, hv in enumerate(headers, 1):
    c = ws.cell(row=1, column=ci, value=hv)
    c.fill = HEADER_FILL; c.font = HDR_FONT
    c.alignment = C; c.border = BORDER

# ── Defect data ────────────────────────────────────────────────────────────────
# (Sl, Submitted By, Date, Description, Sprint, Assigned To,
#  Type, Action Taken, Action Date, Status, Remarks)
defects = [
    (1, "Shubham", "2025-01-10",
     "Bandit linter returning exit code 1 even on syntactically valid Python files with zero security findings, causing the linter pipeline to treat valid submissions as errors.",
     "Sprint 1", "Shubham", "Logical",
     "Added exit-code tolerance in run_bandit(): treat both exit 0 and exit 1 as valid Bandit runs. Exit 2+ still treated as a hard error. Updated linters.py subprocess handler.",
     "2025-01-10", "Closed",
     "Bandit uses exit code 1 to indicate findings exist (not a tool failure). Documented in DEBUGGING_SESSION_NOTES.md Section 1."),

    (2, "Shubham", "2025-01-11",
     "javalang library not listed in pyproject.toml dependencies, causing ImportError at runtime when the Java syntax gatekeeper attempted to parse submitted Java code.",
     "Sprint 1", "Shubham", "Maintainability",
     "Added javalang to [project.dependencies] in pyproject.toml. Ran 'uv sync' to reinstall all dependencies. Verified import resolves correctly at application startup.",
     "2025-01-11", "Closed",
     "Dependency omission discovered during Sprint 1 integration testing. Added version pin javalang>=0.13.0."),

    (3, "Shubham", "2025-01-14",
     "Pylint subprocess failing silently when the temp file path contained spaces or special characters, returning empty JSON output instead of a diagnostic error.",
     "Sprint 1", "Shubham", "Logical",
     "Switched from shell=True subprocess invocation to list-form asyncio.create_subprocess_exec(), which handles paths with spaces correctly without shell interpretation.",
     "2025-01-14", "Closed",
     "Shell=True with string interpolation is also a CWE-78 risk. The fix simultaneously resolved a latent command injection vector. Documented in DEBUGGING_SESSION_NOTES.md Section 3."),

    (4, "Shubham", "2025-01-16",
     "Integration tests failing with AttributeError: module 'app.cache' has no attribute 'redis_cache' after the cache module was refactored to consolidate Redis client management into app/cache.py.",
     "Sprint 1", "Shubham", "Maintainability",
     "Diagnosed that legacy tests were patching internal module internals (app.cache.redis_cache.get_redis_client) rather than the boundary consumer. Realigned all mocks to @patch('app.api.routes.submit.get_redis_client').",
     "2025-01-16", "Closed",
     "Mock patching should always target the import site, not the definition site. Documented as a testing anti-pattern in DEBUGGING_SESSION_NOTES.md Section 8."),

    (5, "Shubham", "2025-01-22",
     "Security Vulnerability Agent returning 'No security vulnerabilities detected!' with a perfect 100/100 score for Python code containing clear SQL Injection (CWE-89) and OS Command Injection (CWE-78) patterns.",
     "Sprint 2", "Shubham", "Logical",
     "Root cause: Gemini content safety filter refused to analyze vulnerability-dense code, returning plaintext refusal instead of JSON. Implemented QA Compliance Auditor prompt framing in security_vuln.py and added deterministic Bandit static fallback via _extract_static_fallbacks().",
     "2025-01-23", "Closed",
     "Documented in DEBUGGING_SESSION_NOTES.md Section 9. LangSmith trace run-019fc6d7 confirmed the refusal. Fallback now guarantees zero vulnerability loss."),

    (6, "Shubham", "2025-01-24",
     "LangGraph pipeline executing agents sequentially, resulting in ~77 seconds total analysis time per submission. Code Analysis and Security agents were waiting for each other unnecessarily.",
     "Sprint 2", "Shubham", "Logical",
     "Redesigned app/agents/graph.py from a linear chain to a parallel fan-out StateGraph. Stage 1 (run_linters) now forks into code_analysis and security_vuln simultaneously. Both converge at sync_findings_node. Reduced pipeline latency by ~50%.",
     "2025-01-25", "Closed",
     "Documented in DEBUGGING_SESSION_NOTES.md Section 10. Verification: LangSmith traces confirmed parallel execution timing."),

    (7, "Shubham", "2025-01-27",
     "Remediation Agent occasionally returning truncated or malformed JSON responses, causing the Pydantic model parser to raise ValidationError and drop all remediation suggestions.",
     "Sprint 2", "Shubham", "Logical",
     "Added 2-attempt retry loop with exponential backoff in run_remediation(). Improved JSON extraction using regex to isolate the first valid JSON object block in the LLM output. Added explicit fallback RemediationResult on second failure.",
     "2025-01-27", "Closed",
     "LLM JSON truncation is a known issue with large outputs. Retry + JSON extraction regex is a robust mitigation pattern."),

    (8, "Shubham", "2025-01-30",
     "Logfire metrics export silently timing out (read timeout 9.9s) causing WARNING log spam in the Celery worker console without impacting actual analysis results.",
     "Sprint 2", "Shubham", "Logical",
     "Configured Logfire async export with non-blocking flush queue. Increased export timeout threshold. Added graceful exception suppression so timeout warnings do not propagate to analysis pipeline.",
     "2025-01-31", "Closed",
     "Intermittent network latency to logfire-us.pydantic.dev. Non-critical — observability telemetry, not pipeline data. Documented in DEBUGGING_SESSION_NOTES.md."),

    (9, "Shubham", "2025-02-01",
     "Security agent failing to detect vulnerabilities in Java code containing JDBC SQL Injection and Path Traversal patterns. LangSmith trace (run-019fc716) showed model refusal for Java security analysis.",
     "Sprint 2", "Shubham", "Logical",
     "Two root causes fixed: (1) Extended _extract_static_fallbacks() to process Java linter heuristics output in addition to Python Bandit results. (2) Strengthened QA Auditor framing to explicitly include Java code review context. Updated fallback IDs to use 'java-stat-NNN' prefix.",
     "2025-02-02", "Closed",
     "Documented in DEBUGGING_SESSION_NOTES.md Section 11. LangSmith trace confirmed fix: Java findings now appear in dashboard."),

    (10, "Shubham", "2025-02-05",
     "re.compile() raising re.error: global flags not at the start of the expression when Java regex patterns used (?i) inline flag inside pattern strings passed to _re.IGNORECASE-compiled expressions.",
     "Sprint 2", "Shubham", "Logical",
     "Replaced all inline (?i) flag syntax inside pattern strings with the _re.IGNORECASE | _re.MULTILINE flags argument in the _re.compile() call. Verified all 14 Java regex patterns compile cleanly.",
     "2025-02-05", "Closed",
     "Python re module does not allow global flags like (?i) after the start of a pattern when using re.compile() with separate flags argument. Fixed in app/linters.py."),

    (11, "Shubham", "2025-02-10",
     "Intent Guardrail validate_intent() incorrectly rejecting valid short Python snippets (e.g., a 2-line function) as non-code input, blocking legitimate submissions from reaching the analysis pipeline.",
     "Sprint 3", "Shubham", "Logical",
     "Tuned the LLM classification prompt to provide explicit examples of short but valid code. Adjusted the minimum length pre-filter threshold from 10 to 5 characters to reduce false rejections. Added test case for short valid Python functions in test_guardrails.py.",
     "2025-02-10", "Closed",
     "Documented in DEBUGGING_SESSION_NOTES.md Section 9. Guardrail now correctly accepts minimal valid code snippets."),

    (12, "Shubham", "2025-02-12",
     "Semgrep first-run causing a 15–20 second cold-start delay when the p/java rulepack was not yet cached locally, causing Celery task timeout on the first Java submission after worker restart.",
     "Sprint 3", "Shubham", "Logical",
     "Semgrep rules are now automatically cached to disk after the first run. Added --timeout 30 flag to run_semgrep_java() subprocess to prevent unbounded cold-start hangs. Documentation updated to advise running one warm-up Java submission after Celery worker restart.",
     "2025-02-13", "Closed",
     "Subsequent runs are near-instant (<2s) as rules are served from local cache. Noted in UI_TEST_CASES.md."),

    (13, "Shubham", "2025-02-14",
     "FAISS RAG vector index not persisting between Celery worker restarts, causing the retriever to return empty context on every new worker process despite the index being built successfully on first startup.",
     "Sprint 3", "Shubham", "Maintainability",
     "Added index serialization to disk (index.faiss + index.pkl) on build completion in app/services/rag/index.py. Retriever now attempts to load from disk on startup before rebuilding. Added existence check before re-indexing.",
     "2025-02-15", "Closed",
     "In-memory FAISS indexes do not survive process restarts. Persistent serialization is required for production reliability."),

    (14, "Shubham", "2025-02-17",
     "pytest integration tests (test_submit_api.py) failing with 422 responses on valid submissions after the Semgrep upgrade changed the Java linter output key from 'pmd' to 'semgrep', breaking the test assertion.",
     "Sprint 3", "Shubham", "User Interface",
     "Updated test_run_java_linters() assertion in tests/unit/test_linters.py to validate the new 'semgrep' and 'heuristics' output keys instead of the legacy 'pmd' key. All 49 tests passing after fix.",
     "2025-02-17", "Closed",
     "Breaking contract change when upgrading from regex heuristics to Semgrep. Test assertions must be updated to reflect new output schema whenever linter architecture changes."),

    (15, "Shubham", "2025-02-19",
     "Chat Graph MemorySaver losing conversational session context after Redis flushall was executed (redis-cli flushall), causing follow-up questions in the Chat Tab to lose reference to the prior code review result.",
     "Sprint 3", "Shubham", "Logical",
     "Scoped MemorySaver checkpoint storage to in-process Python memory (not Redis) to isolate chat session state from the main analysis cache. Chat context is now preserved for the lifetime of the Celery worker process regardless of Redis cache flushes.",
     "2025-02-19", "Closed",
     "Redis flush is a maintenance operation that should not break user chat sessions. In-process MemorySaver is the correct scope for single-user conversation threads."),
]

# ── Write rows ─────────────────────────────────────────────────────────────────
for ri, row in enumerate(defects, 2):
    ws.row_dimensions[ri].height = 70
    rf = ALT1 if ri % 2 == 0 else ALT2
    (sl, sub_by, sub_date, desc, sprint, assigned,
     defect_type, action, action_date, status, remarks) = row

    vals = [sl, sub_by, sub_date, desc, sprint, assigned,
            defect_type, action, action_date, status, remarks]

    for ci, val in enumerate(vals, 1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.border = BORDER
        if ci == 10:  # Status column
            c.fill = CLOSED_FILL if status == "Closed" else OPEN_FILL
            c.font = CLOSED_FNT if status == "Closed" else OPEN_FNT
            c.alignment = C
        elif ci == 7:  # Type of Defect
            type_colors = {
                "Logical":         "FFF3E0",
                "User Interface":  "E3F2FD",
                "Maintainability": "F3E5F5",
                "Standards":       "E8EAF6",
                "Others":          "EFEBE9",
            }
            c.fill = PatternFill("solid", fgColor=type_colors.get(val, "FFFFFF"))
            c.font = BOLD_FONT; c.alignment = C
        elif ci in (1, 2, 3, 5, 6, 9):
            c.fill = rf; c.font = BODY_FONT; c.alignment = C
        else:
            c.fill = rf; c.font = BODY_FONT; c.alignment = L

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:K{len(defects)+1}"

out = "/Users/arous/Desktop/AI-Code-Review-Security-Analysis-Agent-Group2/docs/Defect_Tracker_Filled_v0.1.xlsx"
wb.save(out)
print(f"Saved -> {out}")
print(f"  Total defects logged: {len(defects)}")
closed = sum(1 for d in defects if d[9] == "Closed")
print(f"  Closed: {closed}  |  Open: {len(defects) - closed}")
