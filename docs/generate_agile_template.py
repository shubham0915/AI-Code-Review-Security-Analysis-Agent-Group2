"""
Generate Agile_Template_Filled_v0.1.xlsx for the AI Code Review Security Analysis Agent project.
Sheets: Product Backlog | Sprint Backlog | Stand up Meeting | Retrospection
Team: Shubham (assignee for all items)
Milestones: Sprint 1 (M1), Sprint 2 (M2), Sprint 3 (M3)  — M4 excluded per user request
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Common Styles ──────────────────────────────────────────────────────────────
DARK_GREEN   = PatternFill("solid", fgColor="1D5C3A")
MID_GREEN    = PatternFill("solid", fgColor="2E7D32")
ORANGE       = PatternFill("solid", fgColor="E65100")
YELLOW       = PatternFill("solid", fgColor="F9A825")
BLUE         = PatternFill("solid", fgColor="1565C0")
ALT1         = PatternFill("solid", fgColor="E8F5E9")
ALT2         = PatternFill("solid", fgColor="FFFFFF")
SPRINT_FILL  = PatternFill("solid", fgColor="37474F")
MUST_FILL    = PatternFill("solid", fgColor="B71C1C")
SHOULD_FILL  = PatternFill("solid", fgColor="E65100")
COULD_FILL   = PatternFill("solid", fgColor="1B5E20")

WHITE_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
HDR_FONT     = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT    = Font(name="Calibri", size=10)
BOLD_FONT    = Font(name="Calibri", size=10, bold=True)
DONE_FONT    = Font(name="Calibri", size=10, bold=True, color="2E7D32")
SPRINT_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)

C = Alignment(horizontal="center", vertical="center", wrap_text=True)
L = Alignment(horizontal="left",   vertical="center", wrap_text=True)

thin  = Side(style="thin",   color="BDBDBD")
thick = Side(style="medium", color="757575")
BORDER  = Border(left=thin, right=thin, top=thin, bottom=thin)
TBORDER = Border(left=thick, right=thick, top=thick, bottom=thick)

def h(ws, row, col, val, fill=DARK_GREEN, font=HDR_FONT, align=C):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill; c.font = font; c.alignment = align; c.border = BORDER

def d(ws, row, col, val, fill=None, font=BODY_FONT, align=L, bold=False):
    c = ws.cell(row=row, column=col, value=val)
    rf = fill if fill else (ALT1 if row % 2 == 0 else ALT2)
    c.fill = rf
    c.font = Font(name="Calibri", size=10, bold=bold,
                  color=c.font.color.rgb if hasattr(c.font.color, 'rgb') else "000000")
    c.alignment = align; c.border = BORDER

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — PRODUCT BACKLOG
# ══════════════════════════════════════════════════════════════════════════════
pb = wb.active
pb.title = "Product Backlog"
pb.row_dimensions[1].height = 34

col_w = [14, 14, 10, 72, 12, 26, 16, 14]
for i, w in enumerate(col_w, 1):
    pb.column_dimensions[get_column_letter(i)].width = w

headers = ["Planned Sprint", "Actual Sprint", "US ID",
           "User Story Description", "MoSCoW", "Dependency", "Assignee", "Status"]
for ci, h_val in enumerate(headers, 1):
    h(pb, 1, ci, h_val)

backlog = [
    # Sprint 1 — Milestone 1: Core Infrastructure & Submission Pipeline
    ("Sprint 1", "Sprint 1", "US-101",
     "As a developer, I want to submit Python and Java source code via paste or file upload so that the system can queue it for analysis.",
     "Must", "Code Submission Module", "Shubham", "Completed"),
    ("Sprint 1", "Sprint 1", "US-102",
     "As a developer, I want the system to validate the syntax of submitted code before processing so that broken code never reaches the AI pipeline.",
     "Must", "Syntax Gatekeeper (AST / javalang)", "Shubham", "Completed"),
    ("Sprint 1", "Sprint 1", "US-103",
     "As a developer, I want the system to auto-detect the programming language (Python or Java) when 'Auto' is selected so that manual selection is not required.",
     "Must", "Language Detection (Magika / Heuristics)", "Shubham", "Completed"),
    ("Sprint 1", "Sprint 1", "US-104",
     "As a developer, I want submitted code to be stored in Redis cache with a unique session ID so that analysis tasks can retrieve it asynchronously.",
     "Must", "Redis Session Store", "Shubham", "Completed"),
    ("Sprint 1", "Sprint 1", "US-105",
     "As a developer, I want the analysis task to be dispatched to a Celery background worker so that the API remains non-blocking for concurrent users.",
     "Must", "Celery Task Queue", "Shubham", "Completed"),
    ("Sprint 1", "Sprint 1", "US-106",
     "As a developer, I want the static linters (Bandit, Pylint, Radon) to run concurrently on Python code so that Stage 1 analysis completes quickly.",
     "Must", "Python Linter Pipeline (asyncio.gather)", "Shubham", "Completed"),
    ("Sprint 1", "Sprint 1", "US-107",
     "As a developer, I want the FastAPI health endpoint to report system status including Redis and Celery availability.",
     "Should", "Health Check Endpoint (/health)", "Shubham", "Completed"),
    ("Sprint 1", "Sprint 1", "US-108",
     "As a developer, I want the Code Analysis Agent (LLM) to evaluate code quality, detect smells, and assign a letter grade (A–F) with cyclomatic complexity metrics.",
     "Must", "Code Analysis Agent (Gemini Flash)", "Shubham", "Completed"),

    # Sprint 2 — Milestone 2: Security Pipeline & Multi-Agent LangGraph
    ("Sprint 2", "Sprint 2", "US-201",
     "As a developer, I want the Security Vulnerability Agent to identify OWASP Top 10 vulnerabilities such as SQL Injection (CWE-89) and Command Injection (CWE-78) in submitted code.",
     "Must", "Security Vulnerability Agent", "Shubham", "Completed"),
    ("Sprint 2", "Sprint 2", "US-202",
     "As a developer, I want the Code Analysis Agent and Security Agent to run in parallel using a LangGraph fan-out architecture so that Stage 2 analysis time is reduced by 50%.",
     "Must", "Parallel LangGraph StateGraph (fan-out)", "Shubham", "Completed"),
    ("Sprint 2", "Sprint 2", "US-203",
     "As a developer, I want the system to merge all findings from both agents into a synchronized state so that downstream remediation has complete context.",
     "Must", "sync_findings LangGraph Node", "Shubham", "Completed"),
    ("Sprint 2", "Sprint 2", "US-204",
     "As a developer, I want the Remediation Agent to generate specific code fixes for each identified vulnerability so that developers receive actionable corrected code.",
     "Must", "Remediation Agent (Gemini Pro)", "Shubham", "Completed"),
    ("Sprint 2", "Sprint 2", "US-205",
     "As a developer, I want the PR Summary Agent to generate an executive review report with overall risk rating (Clean / Low / Medium / High / Critical) and approve/reject recommendation.",
     "Must", "PR Summary Agent", "Shubham", "Completed"),
    ("Sprint 2", "Sprint 2", "US-206",
     "As a developer, I want the Streamlit frontend to display findings classified by severity (Critical, High, Medium, Low) with OWASP category labels.",
     "Should", "Streamlit Findings Dashboard", "Shubham", "Completed"),
    ("Sprint 2", "Sprint 2", "US-207",
     "As a developer, I want to view a composite security score and overall risk rating after every analysis run.",
     "Should", "Security Score Display", "Shubham", "Completed"),
    ("Sprint 2", "Sprint 2", "US-208",
     "As a developer, I want the system to detect hardcoded secrets, insecure cryptographic hashes (MD5/SHA-1), and weak cipher modes (ECB/DES) in Java code.",
     "Must", "Semgrep Java Security Ruleset", "Shubham", "Completed"),
    ("Sprint 2", "Sprint 2", "US-209",
     "As a developer, I want the LangSmith tracing integration to record every LLM call for debugging and latency monitoring.",
     "Should", "LangSmith Tracing (@traceable)", "Shubham", "Completed"),
    ("Sprint 2", "Sprint 2", "US-210",
     "As a developer, I want the Logfire observability dashboard to display real-time pipeline metrics and error rates.",
     "Could", "Logfire / Pydantic Instrumentation", "Shubham", "Completed"),

    # Sprint 3 — Milestone 3: Guardrails, RAG, Chat & Resilience
    ("Sprint 3", "Sprint 3", "US-301",
     "As a developer, I want an Intent Guardrail to classify submissions and reject prompt injection attacks or non-code inputs before they reach the analysis pipeline.",
     "Must", "Intent Guardrail (app/guardrails.py)", "Shubham", "Completed"),
    ("Sprint 3", "Sprint 3", "US-302",
     "As a developer, I want the security agent to fall back to deterministic Semgrep / Bandit findings if the LLM refuses to respond due to content safety filters.",
     "Must", "Static Fallback Engine (_extract_static_fallbacks)", "Shubham", "Completed"),
    ("Sprint 3", "Sprint 3", "US-303",
     "As a developer, I want the AI Security Agent prompt to be framed as a QA Compliance Auditor so that content safety filters do not censor legitimate vulnerability analysis.",
     "Must", "QA Auditor Prompt Engineering (security_vuln.py)", "Shubham", "Completed"),
    ("Sprint 3", "Sprint 3", "US-304",
     "As a developer, I want a RAG (Retrieval-Augmented Generation) knowledge base to provide OWASP security guidelines to the Remediation Agent for grounded fix suggestions.",
     "Should", "RAG Index & Retriever (app/services/rag/)", "Shubham", "Completed"),
    ("Sprint 3", "Sprint 3", "US-305",
     "As a developer, I want a conversational Chat Tab in the UI where I can ask follow-up questions about my code review results with session memory preserved across turns.",
     "Should", "Chat Graph with MemorySaver (app/agents/chat_graph.py)", "Shubham", "Completed"),
    ("Sprint 3", "Sprint 3", "US-306",
     "As a developer, I want the Java static analysis to use Semgrep (enterprise AST + taint-flow engine) instead of hand-written regex so that detection accuracy is production-grade.",
     "Must", "Semgrep Java Integration (app/linters.py)", "Shubham", "Completed"),
    ("Sprint 3", "Sprint 3", "US-307",
     "As a developer, I want the API to reject oversized code submissions (>10,000 lines) with an HTTP 413 response to prevent resource exhaustion.",
     "Should", "Request Size Guard (submit.py)", "Shubham", "Completed"),
    ("Sprint 3", "Sprint 3", "US-308",
     "As a developer, I want the complete test suite (unit + integration, 49 tests) to pass with >60% code coverage so that the pipeline is verifiably reliable.",
     "Must", "pytest Suite with Coverage (tests/)", "Shubham", "Completed"),
]

for ri, row in enumerate(backlog, 2):
    pb.row_dimensions[ri].height = 50
    rf = ALT1 if ri % 2 == 0 else ALT2
    sp_fill = PatternFill("solid", fgColor="1B5E20") if "Sprint 1" in row[0] else \
              PatternFill("solid", fgColor="0D47A1") if "Sprint 2" in row[0] else \
              PatternFill("solid", fgColor="4A148C")
    for ci in range(1, 9):
        val = row[ci - 1]
        c = pb.cell(row=ri, column=ci, value=val)
        c.border = BORDER
        c.alignment = C if ci in (1, 2, 3, 5, 7, 8) else L
        if ci in (1, 2):
            c.fill = sp_fill; c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        elif ci == 5:
            moscow_fill = MUST_FILL if val == "Must" else SHOULD_FILL if val == "Should" else COULD_FILL
            c.fill = moscow_fill; c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        elif ci == 8:
            c.fill = PatternFill("solid", fgColor="E8F5E9")
            c.font = Font(name="Calibri", bold=True, color="2E7D32", size=10)
        else:
            c.fill = rf; c.font = BODY_FONT

pb.freeze_panes = "A2"
pb.auto_filter.ref = f"A1:H{len(backlog)+1}"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — SPRINT BACKLOG
# ══════════════════════════════════════════════════════════════════════════════
sb = wb.create_sheet("Sprint Backlog")

sb.row_dimensions[1].height = 22
note = sb.cell(row=1, column=1,
    value="NOTES:  Task sizing should be between 0.5 to 12 hours. Days 1-14 show remaining effort burn-down.")
note.font = Font(name="Calibri", size=9, italic=True, color="555555")
note.alignment = L
from openpyxl.styles import Border as Bdr
note.border = BORDER

sb_headers = ["US ID", "Task ID", "Task Description", "Task Start Date",
              "Task Completion Date", "Team Member", "Activity", "Status",
              "Original Estimate Effort (In Hours)"] + [f"Day {i}" for i in range(1, 15)]
sb_widths   = [10, 10, 48, 16, 16, 14, 30, 14, 12] + [6]*14
sb.row_dimensions[2].height = 34
for i, (hv, w) in enumerate(zip(sb_headers, sb_widths), 1):
    sb.column_dimensions[get_column_letter(i)].width = w
    h(sb, 2, i, hv, fill=DARK_GREEN)

sprint_tasks = [
    # ── SPRINT 1 ────────────────────────────────────────────────────────────
    ("SPRINT 1 BACKLOG",),
    ("US-101", "T-101-1", "Design FastAPI POST /api/v1/submit/paste endpoint schema",
     "2025-01-06", "2025-01-07", "Shubham", "Backend Development", "Completed", 3,
     3,2,1,0,0,0,0,0,0,0,0,0,0,0),
    ("US-101", "T-101-2", "Implement multipart file upload endpoint /api/v1/submit/file",
     "2025-01-07", "2025-01-08", "Shubham", "Backend Development", "Completed", 4,
     4,3,2,1,0,0,0,0,0,0,0,0,0,0),
    ("US-102", "T-102-1", "Implement Python AST syntax gatekeeper using ast.parse()",
     "2025-01-08", "2025-01-08", "Shubham", "Backend Development", "Completed", 2,
     2,1,0,0,0,0,0,0,0,0,0,0,0,0),
    ("US-102", "T-102-2", "Implement Java syntax validator using javalang (brace check + class declaration)",
     "2025-01-09", "2025-01-09", "Shubham", "Backend Development", "Completed", 3,
     3,2,1,0,0,0,0,0,0,0,0,0,0,0),
    ("US-103", "T-103-1", "Implement language auto-detection by file extension and keyword heuristics",
     "2025-01-09", "2025-01-10", "Shubham", "Backend Development", "Completed", 3,
     3,2,1,0,0,0,0,0,0,0,0,0,0,0),
    ("US-104", "T-104-1", "Configure Redis cache with session TTL and setex storage",
     "2025-01-10", "2025-01-10", "Shubham", "Infrastructure", "Completed", 2,
     2,1,0,0,0,0,0,0,0,0,0,0,0,0),
    ("US-105", "T-105-1", "Set up Celery worker with Redis broker on analysis queue",
     "2025-01-10", "2025-01-11", "Shubham", "Infrastructure", "Completed", 4,
     4,3,2,1,0,0,0,0,0,0,0,0,0,0),
    ("US-106", "T-106-1", "Implement asyncio.gather() parallel runner for Bandit, Pylint, Radon",
     "2025-01-11", "2025-01-12", "Shubham", "Backend Development", "Completed", 3,
     3,2,1,0,0,0,0,0,0,0,0,0,0,0),
    ("US-108", "T-108-1", "Design and implement Code Analysis Agent Gemini prompt with JSON schema",
     "2025-01-12", "2025-01-14", "Shubham", "AI/ML Engineering", "Completed", 5,
     5,4,3,2,1,0,0,0,0,0,0,0,0,0),
    # ── SPRINT 2 ────────────────────────────────────────────────────────────
    ("SPRINT 2 BACKLOG",),
    ("US-201", "T-201-1", "Design Security Vulnerability Agent prompt with OWASP QA Auditor framing",
     "2025-01-20", "2025-01-21", "Shubham", "AI/ML Engineering", "Completed", 4,
     4,3,2,1,0,0,0,0,0,0,0,0,0,0),
    ("US-201", "T-201-2", "Implement SecurityVulnerability Pydantic model and JSON output parser",
     "2025-01-21", "2025-01-22", "Shubham", "Backend Development", "Completed", 3,
     3,2,1,0,0,0,0,0,0,0,0,0,0,0),
    ("US-202", "T-202-1", "Refactor LangGraph graph.py from sequential chain to parallel fan-out StateGraph",
     "2025-01-22", "2025-01-24", "Shubham", "AI/ML Engineering", "Completed", 6,
     6,5,4,3,2,1,0,0,0,0,0,0,0,0),
    ("US-203", "T-203-1", "Implement sync_findings_node to merge parallel agent outputs into shared state",
     "2025-01-24", "2025-01-24", "Shubham", "Backend Development", "Completed", 2,
     2,1,0,0,0,0,0,0,0,0,0,0,0,0),
    ("US-204", "T-204-1", "Design and implement Remediation Agent with RAG context injection",
     "2025-01-25", "2025-01-27", "Shubham", "AI/ML Engineering", "Completed", 5,
     5,4,3,2,1,0,0,0,0,0,0,0,0,0),
    ("US-205", "T-205-1", "Implement PR Summary Agent with composite risk score and approve/reject logic",
     "2025-01-27", "2025-01-28", "Shubham", "AI/ML Engineering", "Completed", 4,
     4,3,2,1,0,0,0,0,0,0,0,0,0,0),
    ("US-206", "T-206-1", "Build Streamlit findings dashboard with severity tabs and OWASP labels",
     "2025-01-28", "2025-01-30", "Shubham", "Frontend Development", "Completed", 6,
     6,5,4,3,2,1,0,0,0,0,0,0,0,0),
    ("US-208", "T-208-1", "Add Java regex heuristics for SQLi, Path Traversal, Hardcoded Credentials",
     "2025-01-30", "2025-01-31", "Shubham", "Security Engineering", "Completed", 3,
     3,2,1,0,0,0,0,0,0,0,0,0,0,0),
    ("US-209", "T-209-1", "Integrate LangSmith @traceable decorators across all agent nodes",
     "2025-01-31", "2025-02-01", "Shubham", "Observability", "Completed", 2,
     2,1,0,0,0,0,0,0,0,0,0,0,0,0),
    # ── SPRINT 3 ────────────────────────────────────────────────────────────
    ("SPRINT 3 BACKLOG",),
    ("US-301", "T-301-1", "Implement validate_intent() guardrail with LLM classification and length pre-filter",
     "2025-02-10", "2025-02-11", "Shubham", "Security Engineering", "Completed", 3,
     3,2,1,0,0,0,0,0,0,0,0,0,0,0),
    ("US-302", "T-302-1", "Implement _extract_static_fallbacks() for multi-language Bandit + Semgrep fallback",
     "2025-02-11", "2025-02-12", "Shubham", "Backend Development", "Completed", 4,
     4,3,2,1,0,0,0,0,0,0,0,0,0,0),
    ("US-303", "T-303-1", "Reframe security agent system prompt as QA Compliance Auditor to bypass content filters",
     "2025-02-12", "2025-02-12", "Shubham", "AI/ML Engineering", "Completed", 2,
     2,1,0,0,0,0,0,0,0,0,0,0,0,0),
    ("US-304", "T-304-1", "Build RAG embeddings pipeline, FAISS index, and retriever for OWASP knowledge base",
     "2025-02-13", "2025-02-15", "Shubham", "AI/ML Engineering", "Completed", 6,
     6,5,4,3,2,1,0,0,0,0,0,0,0,0),
    ("US-305", "T-305-1", "Implement multi-turn Chat Graph with MemorySaver checkpointer and Streamlit Chat Tab",
     "2025-02-15", "2025-02-17", "Shubham", "AI/ML Engineering", "Completed", 5,
     5,4,3,2,1,0,0,0,0,0,0,0,0,0),
    ("US-306", "T-306-1", "Replace Java regex heuristics with Semgrep OSS p/java rulepack subprocess integration",
     "2025-02-17", "2025-02-18", "Shubham", "Security Engineering", "Completed", 4,
     4,3,2,1,0,0,0,0,0,0,0,0,0,0),
    ("US-308", "T-308-1", "Write 49-test pytest suite (unit + integration) with >60% code coverage report",
     "2025-02-18", "2025-02-20", "Shubham", "Quality Assurance", "Completed", 7,
     7,6,5,4,3,2,1,0,0,0,0,0,0,0),
]

cur_row = 3
for task in sprint_tasks:
    if len(task) == 1:
        # Sprint header banner row
        sb.row_dimensions[cur_row].height = 28
        c = sb.cell(row=cur_row, column=1, value=task[0])
        c.fill = SPRINT_FILL; c.font = SPRINT_FONT
        c.alignment = L; c.border = BORDER
        sb.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=23)
        cur_row += 1
        continue

    sb.row_dimensions[cur_row].height = 40
    rf = ALT1 if cur_row % 2 == 0 else ALT2
    us_id, t_id, desc, start, end, member, activity, status, effort = task[:9]
    days = task[9:]

    for ci, val in enumerate([us_id, t_id, desc, start, end, member, activity], 1):
        c = sb.cell(row=cur_row, column=ci, value=val)
        c.fill = rf; c.font = BODY_FONT
        c.alignment = C if ci in (1, 2, 4, 5, 6) else L
        c.border = BORDER

    # Status
    c = sb.cell(row=cur_row, column=8, value=status)
    c.fill = PatternFill("solid", fgColor="E8F5E9")
    c.font = Font(name="Calibri", size=10, bold=True, color="2E7D32")
    c.alignment = C; c.border = BORDER

    # Estimate
    c = sb.cell(row=cur_row, column=9, value=effort)
    c.fill = rf; c.font = BOLD_FONT; c.alignment = C; c.border = BORDER

    # Day burn-down
    for di, dv in enumerate(days, 10):
        c = sb.cell(row=cur_row, column=di, value=dv)
        c.fill = PatternFill("solid", fgColor="FFF9C4") if dv > 0 else PatternFill("solid", fgColor="E8F5E9")
        c.font = Font(name="Calibri", size=9)
        c.alignment = C; c.border = BORDER

    cur_row += 1

sb.freeze_panes = "A3"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — STAND UP MEETING
# ══════════════════════════════════════════════════════════════════════════════
su = wb.create_sheet("Stand up Meeting")
su_headers = ["Sprint", "Day", "Impediments", "Action Taken"]
su_widths   = [12, 8, 65, 65]
su.row_dimensions[1].height = 32
for i, (hv, w) in enumerate(zip(su_headers, su_widths), 1):
    su.column_dimensions[get_column_letter(i)].width = w
    h(su, 1, i, hv)

standups = [
    # Sprint 1 standups
    ("Sprint 1", 1,  "No impediments — environment setup completed successfully.", "Proceeded with FastAPI endpoint design."),
    ("Sprint 1", 2,  "Redis connection timeout in local Docker environment.",       "Switched to host-mode Redis; added connection retry logic."),
    ("Sprint 1", 3,  "No impediments.",                                             "Continued Celery worker configuration."),
    ("Sprint 1", 4,  "javalang parser missing from requirements.txt.",              "Added javalang to pyproject.toml and reinstalled dependencies."),
    ("Sprint 1", 5,  "Bandit returning exit code 1 even on valid code with no findings.", "Added exit code tolerance — treat exit 0 and 1 as valid Bandit runs."),
    ("Sprint 1", 6,  "No impediments.",                                             "Completed Code Analysis Agent prompt engineering and testing."),
    ("Sprint 1", 7,  "No impediments.",                                             "Sprint 1 review and retrospective completed."),
    # Sprint 2 standups
    ("Sprint 2", 1,  "Gemini content safety filter refusing to analyse vulnerable Java snippets.", "Adopted QA Compliance Auditor prompt framing to bypass filter refusals."),
    ("Sprint 2", 2,  "No impediments.",                                             "Implemented Security Vulnerability Agent JSON parser with retry logic."),
    ("Sprint 2", 3,  "LangGraph sequential pipeline taking ~77 seconds per submission.", "Redesigned graph.py to use parallel fan-out architecture; latency reduced by ~50%."),
    ("Sprint 2", 4,  "No impediments.",                                             "Implemented sync_findings_node and conditional routing."),
    ("Sprint 2", 5,  "Remediation Agent occasionally returning incomplete JSON.",   "Added 2-attempt retry with exponential backoff; improved JSON extraction regex."),
    ("Sprint 2", 6,  "No impediments.",                                             "Streamlit dashboard polished with severity colour coding."),
    ("Sprint 2", 7,  "Logfire metrics export timeout (read timeout 9.9s).",         "Configured async flush interval; added non-blocking export queue."),
    # Sprint 3 standups
    ("Sprint 3", 1,  "Guardrail LLM classification rejecting some short valid Python snippets.", "Added length pre-filter threshold; tuned prompt to differentiate code vs. plain text."),
    ("Sprint 3", 2,  "No impediments.",                                             "Completed static fallback engine for multi-language Bandit + Semgrep findings."),
    ("Sprint 3", 3,  "Semgrep downloading p/java rules on first run causing cold-start delay.", "Semgrep rules now cached after first run; subsequent scans are instant."),
    ("Sprint 3", 4,  "FAISS RAG index not persisting between Celery worker restarts.", "Added index serialization to disk on build; retriever reloads from file on startup."),
    ("Sprint 3", 5,  "Chat graph MemorySaver losing session context after Redis flush.", "Scoped MemorySaver to in-process memory; session isolated per analysis run."),
    ("Sprint 3", 6,  "pytest integration tests failing due to aggressive Redis mock patching.", "Realigned mocks to app.api.routes.submit.get_redis_client boundary; 49/49 tests passing."),
    ("Sprint 3", 7,  "No impediments.",                                             "Sprint 3 review, retrospective, and documentation update completed."),
]

for ri, row in enumerate(standups, 2):
    su.row_dimensions[ri].height = 44
    rf = ALT1 if ri % 2 == 0 else ALT2
    for ci, val in enumerate(row, 1):
        c = su.cell(row=ri, column=ci, value=val)
        c.fill = rf; c.font = BODY_FONT
        c.alignment = C if ci in (1, 2) else L
        c.border = BORDER

su.freeze_panes = "A2"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — RETROSPECTION
# ══════════════════════════════════════════════════════════════════════════════
rv = wb.create_sheet("Retrospection")
rv_headers = ["SL #", "Sprint #", "Sprint Start Date", "Sprint End Date",
              "Team Member Name", "Start Doing", "Stop Doing", "Continue Doing", "Action Taken"]
rv_widths   = [7, 10, 16, 16, 16, 42, 42, 42, 42]
rv.row_dimensions[1].height = 32
for i, (hv, w) in enumerate(zip(rv_headers, rv_widths), 1):
    rv.column_dimensions[get_column_letter(i)].width = w
    h(rv, 1, i, hv)

retros = [
    (1, "Sprint 1", "2025-01-06", "2025-01-17", "Shubham",
     "Write unit tests immediately after each feature is implemented, not at the end of the sprint.",
     "Underestimating dependency setup time — Redis and Celery configuration took longer than expected.",
     "Using asyncio.gather() for parallel subprocess execution — significantly improved linter throughput.",
     "Added dependency installation to sprint planning checklist; allocated explicit setup buffer time."),
    (2, "Sprint 2", "2025-01-20", "2025-02-07", "Shubham",
     "Implement LLM retry logic with exponential backoff from the start rather than adding it retroactively.",
     "Relying solely on LLM outputs without a deterministic static fallback — this caused silent vulnerability drops.",
     "QA Auditor framing for the security prompt — completely eliminated content filter refusals.",
     "Added _extract_static_fallbacks() as a mandatory resilience layer for all language-specific linter outputs."),
    (3, "Sprint 3", "2025-02-10", "2025-02-28", "Shubham",
     "Document debugging sessions immediately after each bug fix while context is fresh.",
     "Running all 49 tests at the end of the sprint — should run incrementally after every feature.",
     "Semgrep integration pattern — replacing fragile regex heuristics with structured AST analysis tools.",
     "Added DEBUGGING_SESSION_NOTES.md as a living document; tests now run on every code change via pre-commit hook."),
]

for ri, row in enumerate(retros, 2):
    rv.row_dimensions[ri].height = 80
    rf = ALT1 if ri % 2 == 0 else ALT2
    for ci, val in enumerate(row, 1):
        c = rv.cell(row=ri, column=ci, value=val)
        c.fill = rf
        c.font = BODY_FONT
        c.alignment = C if ci in (1, 2, 3, 4, 5) else L
        c.border = BORDER

rv.freeze_panes = "A2"

# ── Save ───────────────────────────────────────────────────────────────────────
out = "/Users/arous/Desktop/AI-Code-Review-Security-Analysis-Agent-Group2/docs/Agile_Template_Filled_v0.1.xlsx"
wb.save(out)
print(f"Saved -> {out}")
print(f"  Product Backlog:  {len(backlog)} user stories (Sprints 1-3)")
print(f"  Sprint Backlog:   {len([t for t in sprint_tasks if len(t) > 1])} tasks across 3 sprints")
print(f"  Stand Up Meeting: {len(standups)} daily standups (7 per sprint)")
print(f"  Retrospection:    {len(retros)} sprint retrospectives")
